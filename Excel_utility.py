########################################################################################################################
# Author : Utkarsh Singh
# Project : Daily Utility
# Module : Excel Utility this is made to add a row or to change the cell value of the row of a particular column
# Last Modified Date : 13 Dec,2020
########################################################################################################################
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from tkinter import *
import os
import MCPI
import openpyxl


class excel_util(Toplevel):
    def __init__(self):
        Toplevel.__init__(self)
        self.geometry("350x250+790+120")
        self.title("Excel Util")
        self.resizable(False, False)
        self.focus()

        # Top frame to open excel file
        self.excel_frame = Frame(self, height=250)
        self.excel_frame.pack(fill='both')

        style = ttk.Style()
        style.configure("Treeview",background='#D3D3D3',foreground='black',rowheight=25, feildbackground="#D3D3D3")

        # Create a treeview
        self.treeview = ttk.Treeview(self.excel_frame)
        self.treescrolly = ttk.Scrollbar(self.excel_frame, orient="vertical",
                                         command=self.treeview.yview)
        self.treescrollx = ttk.Scrollbar(self.excel_frame, orient="horizontal",
                                         command=self.treeview.xview)

        # Add menu
        self.top_menu = Menu(self)
        self.config(menu=self.top_menu)

        # Add menu dropdown
        file_menu = Menu(self.top_menu, tearoff=False)
        self.top_menu.add_cascade(label='Spreadsheets', menu=file_menu)
        file_menu.add_command(label='Open', command=self.file_open)

        # Frame to add a new row in excel
        self.btmFrame = Frame(self, height=500, bg='#58a9b0')
        self.btmFrame.pack(fill="both")

        self.addFrame = LabelFrame(self.btmFrame, text="Add New Row", bg="#58a9b0")
        self.addFrame.place(x=5, y=10, height=70, width=700)

        self.modify_Frame = LabelFrame(self.btmFrame, text="Modify a Cell", bg="#58a9b0")
        self.modify_Frame.place(x=5, y=110, height=180, width=700)
        # Image for successful window
        self.success_image = PhotoImage(file=os.path.join(MCPI.path, 'Images/details.png'))

        # Modify Cell
        self.modify_cell_name_label = ttk.Label(self.modify_Frame, text="Column Number", font=("Times New Roman", 10)
                                                , background="#58a9b0").place(x=5, y=5)
        self.modify_clmn_name_entry = ttk.Entry(self.modify_Frame, width=10)

        self.modify_row_label = ttk.Label(self.modify_Frame, text="Row Number", font=("Times New Roman", 10)
                                          , background="#58a9b0").place(x=5, y=40)
        self.modify_row_name_entry = ttk.Entry(self.modify_Frame, width=10)

        self.modify_cell_value_label = ttk.Label(self.modify_Frame, text="Value", font=("Times New Roman", 10)
                                                 , background="#58a9b0").place(x=5, y=80)
        self.modify_cell_value_entry = ttk.Entry(self.modify_Frame, width=20)

    # Open the excel file in tree-view over the same window and also only after opening the spreadsheet other options on
    # window gets Visible
    def file_open(self):
        self.treeview.delete(*self.treeview.get_children())  # Every time this line will refresh the table
        self.geometry("750x580+790+120")
        filename = filedialog.askopenfilename(
            initialdir=os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop/Screenshot/My Reports'),
            title='Open A File',
            filetype=(("xslx files", "*.xlsx"), ("xsl files", "*.xls"), ("All Files", "*.*"))
        )
        # Read File and Open it in tree-view
        self.table_view(filename)
        try:
            self.title("Excel Util {}".format(filename))

            # Radio Buttons for Adding a defect or Pass Test Case
            option = StringVar()
            radio_success = Radiobutton(self.addFrame, text='Add Success Run', value='Success', variable=option,
                                        bg="#58a9b0", command=lambda: self.add_successrowinexcel(filename))
            radio_success.place(y=5)
            radio_fail = Radiobutton(self.addFrame, text='Add Defect Run', value='Fail', variable=option,
                                     bg="#58a9b0", command=lambda: self.add_defectrowinexcel(filename))
            radio_fail.place(x=205, y=5)

            # Packing Modify options
            self.modify_clmn_name_entry.place(x=150, y=5)
            self.modify_row_name_entry.place(x=150, y=40)
            self.modify_cell_value_entry.place(x=150, y=80)
            modify_btn = ttk.Button(self.modify_Frame, width=14, text="Change",
                                    command=lambda: self.modify_a_cell(filename, self.modify_row_name_entry.get(),
                                                                       self.modify_clmn_name_entry.get(),
                                                                       self.modify_cell_value_entry.get()))
            modify_btn.place(x=50, y=110)



        except:
            pass

    # Function to modify the cell
    def modify_a_cell(self, file, row, clmn, m_value):
        try:
            if row and clmn and m_value != "":
                filename = file
                file = file.split("/")
                xlfile = file.pop()
                file = '/'.join(file)
                os.chdir(file)
                wb = openpyxl.load_workbook(xlfile)
                sheet = wb.active
                row = int(row)
                sheet.cell((row + 1), int(clmn)).value = m_value
                wb.save(xlfile)
                self.table_view(filename)
                self.modify_row_name_entry.delete(0, END)
                self.modify_clmn_name_entry.delete(0, END)
                self.modify_cell_value_entry.delete(0, END)
            else:
                messagebox.showerror("Missing Values", "Please enter values in the fields")
        except Exception as e:
            messagebox.showerror("Invalid Input", f"{e}\nYou could have missed this"
                                                  "\n1.Enter Numeric Values for the column and row field"
                                                  "\n2.May be the file is in open state close it and try again")

    # Update tree-view
    def table_view(self, filename):
        # After the file gets saved present the changed data on the tree-view over the same window
        self.treeview.delete(*self.treeview.get_children())
        global df
        if filename:
            try:
                filename = r"{}".format(filename)
                df = pd.read_excel(filename)
            except:
                messagebox.showerror("Error", "File Cannot be open....Try Again", icon='error')
        elif filename == '':
            messagebox.showerror("Error", "Please Select a file", icon='error')
            self.destroy()
            excel_util()
        try:
            self.treeview.delete(*self.treeview.get_children())
            self.treeview.tag_configure("oddrows", background="white")
            self.treeview.tag_configure("evenrows", background="lightblue")

            self.treeview["column"] = list(df.columns)
            self.treeview["show"] = "headings"

            for column in self.treeview["column"]:
                self.treeview.heading(column, text=column)

            df_rows = df.to_numpy().tolist()

            global count
            count = 0
            for rows in df_rows:
                if count % 2 == 0:
                    self.treeview.insert("", "end", values=rows, iid=count, tags=('oddrows',))
                else:
                    self.treeview.insert("", "end", values=rows, iid=count, tags=('evenrows',))
                count += 1
            self.treeview.configure(xscrollcommand=self.treescrollx.set,
                                    yscrollcommand=self.treescrolly.set)
            self.treescrolly.pack(side="right", fill="y")
            self.treeview.pack()
            self.treescrollx.pack(side="bottom", fill="x")
        except Exception as e:
            messagebox.showerror("Wait!!", e)
            pass

    # If user wants to add a successful run in the excel
    def add_successrowinexcel(self, file):
        file = file.split("/")
        xlfile = file.pop()
        file = '/'.join(file)
        os.chdir(file)
        success_window = Toplevel(self)
        success_window.geometry("500x450+720+120")
        success_window.title("Details of Pass Test Case")

        # Top frame to hold heading and  photo
        success_topframe = Frame(success_window, height=150, bg="white")
        success_topframe.pack(fill="both")
        success_image_label = ttk.Label(success_topframe, image=self.success_image, background="white")
        success_image_label.place(x=30, y=50)
        success_toplabel = ttk.Label(success_topframe, text="Add TC Details", font=("Times New Roman", 15)
                                     , background="white")
        success_toplabel.place(x=190, y=80)

        success_btm_frame = Frame(success_window, height=500, bg="#58a9b0")
        success_btm_frame.pack(fill="both")

        # Bottom frame to hold all input widget
        success_App_name_label = ttk.Label(success_btm_frame, text="App. Name", font=("Times New Roman", 12)
                                           , background="#58a9b0").place(x=10, y=50)
        success_App_name = ttk.Entry(success_btm_frame, width=15)
        success_App_name.focus_force()
        success_App_name.place(x=180, y=50)

        success_Tc_name_label = ttk.Label(success_btm_frame, text="TestCase Name", font=("Times New Roman", 12)
                                          , background="#58a9b0").place(x=10, y=90)
        success_Tc_name = ttk.Entry(success_btm_frame, width=15)
        success_Tc_name.place(x=180, y=90)

        success_date_name_label = ttk.Label(success_btm_frame, text="Date", font=("Times New Roman", 12)
                                            , background="#58a9b0").place(x=10, y=130)
        success_date_name = ttk.Entry(success_btm_frame, width=15)
        success_date_name.place(x=180, y=130)

        success_loc_label = ttk.Label(success_btm_frame, text="Location", font=("Times New Roman", 12)
                                      , background="#58a9b0").place(x=10, y=170)
        success_loc_name = ttk.Entry(success_btm_frame, width=15)
        success_loc_name.place(x=180, y=170)
        success_btn = ttk.Button(success_btm_frame, width=14, text='DONE',
                                 command=lambda: save_xl(xlfile, success_App_name.get(), success_Tc_name.get(),
                                                         success_date_name.get(), success_loc_name.get()))
        success_btn.place(x=120, y=230)

        def save_xl(xlfile, app_name, tc_name, date, loc):
            wb = openpyxl.load_workbook(xlfile)
            sheet = wb.active
            sr_no = sheet.max_row
            sheet.append([sr_no, app_name, tc_name, date, "Pass", loc, ' ', ' ', ' '])
            wb.save(xlfile)
            success_window.destroy()
            messagebox.showinfo("Done!", "Row Added open the file")
            self.table_view(xlfile)

    def add_defectrowinexcel(self, file):
        file = file.split("/")
        xlfile = file.pop()
        file = '/'.join(file)
        os.chdir(file)
        defect_window = Toplevel(self)
        defect_window.geometry("600x610+720+120")
        defect_window.title("Details of Fail Test Case")

        # Top frame to hold heading and  photo
        defect_topframe = Frame(defect_window, height=150, bg="white")
        defect_topframe.pack(fill="both")
        defect_image_label = ttk.Label(defect_topframe, image=self.success_image, background="white")
        defect_image_label.place(x=30, y=50)
        defect_toplabel = ttk.Label(defect_topframe, text="Add TC Details", font=("Times New Roman", 15)
                                    , background="white")
        defect_toplabel.place(x=190, y=80)

        defect_btm_frame = Frame(defect_window, height=600, bg="#782828")
        defect_btm_frame.pack(fill="both")

        # Bottom frame to hold all input widget
        defect_App_name_label = ttk.Label(defect_btm_frame, text="App. Name", font=("Times New Roman", 12)
                                          , background="#782828").place(x=10, y=50)
        defect_App_name = ttk.Entry(defect_btm_frame, width=15)
        defect_App_name.focus_force()
        defect_App_name.place(x=180, y=50)

        defect_Tc_name_label = ttk.Label(defect_btm_frame, text="TestCase Name", font=("Times New Roman", 12)
                                         , background="#782828").place(x=10, y=90)
        defect_Tc_name = ttk.Entry(defect_btm_frame, width=15)
        defect_Tc_name.place(x=180, y=90)

        defect_date_name_label = ttk.Label(defect_btm_frame, text="Date", font=("Times New Roman", 12)
                                           , background="#782828").place(x=10, y=130)
        defect_date_name = ttk.Entry(defect_btm_frame, width=15)
        defect_date_name.place(x=180, y=130)

        defect_loc_label = ttk.Label(defect_btm_frame, text="Location", font=("Times New Roman", 12)
                                     , background="#782828").place(x=10, y=170)
        defect_loc_name = ttk.Entry(defect_btm_frame, width=15)
        defect_loc_name.place(x=180, y=170)

        defect_id_label = ttk.Label(defect_btm_frame, text="Defect Id", font=("Times New Roman", 12)
                                    , background="#782828").place(x=10, y=210)
        defect_id_name = ttk.Entry(defect_btm_frame, width=15)
        defect_id_name.place(x=180, y=210)

        defect_des_label = ttk.Label(defect_btm_frame, text="Short desc.", font=("Times New Roman", 12)
                                     , background="#782828").place(x=10, y=250)
        defect_des_name = ttk.Entry(defect_btm_frame, width=15)
        defect_des_name.place(x=180, y=250)

        defect_scrn_label = ttk.Label(defect_btm_frame, text="Screen Name", font=("Times New Roman", 12)
                                      , background="#782828").place(x=10, y=290)
        defect_scrn_name = ttk.Entry(defect_btm_frame, width=15)
        defect_scrn_name.place(x=180, y=290)

        defect_btn = ttk.Button(defect_btm_frame, width=14, text='DONE',
                                command=lambda: save_xl(xlfile, defect_App_name.get(),
                                                        defect_Tc_name.get(), defect_date_name.get(),
                                                        defect_loc_name.get(), defect_id_name.get(),
                                                        defect_des_name.get(), defect_scrn_name.get()))
        defect_btn.place(x=120, y=330)

        def save_xl(xllfile, app_name, tc_name, date, loc, d_id, d_des, d_sc):
            wb = openpyxl.load_workbook(xllfile)
            sheet = wb.active
            sr_no = sheet.max_row
            sheet.append([sr_no, app_name, tc_name, date, "Pass", loc, d_id, d_des, d_sc])
            wb.save(xlfile)
            defect_window.destroy()
            messagebox.showinfo("Done!", "Row Added open the file")
            self.table_view(xlfile)
